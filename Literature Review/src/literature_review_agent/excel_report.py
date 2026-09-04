"""``Literature_Review_Matrix.xlsx``: the ten-sheet literature-review matrix.

One row per paper in ``Master Matrix``. Formatting follows the rules a reviewer
actually needs: frozen headers, filters, readable widths, wrapped text,
hyperlinks, colour-coded verification status, and no merged cells inside data
tables. Column layout comes from ``config/report_columns.yaml``, and any column
the agent adds at runtime is documented in the ``Data Dictionary`` sheet.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .config import Settings
from .logging_setup import get_logger
from .schemas import (
    CitationAuditRow,
    EvidenceStance,
    GapItem,
    JobConfig,
    LandscapeSummary,
    ModelProfile,
    PaperAnalysis,
    PaperRecord,
)
from .utils import ensure_dir, truncate_text

LOG = get_logger("excel")

# --- styling constants -------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
BODY_FONT = Font(size=10)
LINK_FONT = Font(size=10, color="0563C1", underline="single")
THIN = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

#: Longest string written into any cell; Excel's own limit is 32,767.
MAX_CELL_CHARS = 3000

#: Value written whenever a field was not reported by the source.
NOT_REPORTED = "Information not reported"


@dataclass
class TopicColumn:
    """One agent-added, topic-specific column."""

    header: str
    definition: str
    rationale: str
    values: dict[str, str] = field(default_factory=dict)
    width: int = 30


@dataclass
class MatrixInputs:
    """Everything the workbook builder needs."""

    config: JobConfig
    records: list[PaperRecord]
    analyses: dict[str, PaperAnalysis] = field(default_factory=dict)
    gaps: list[GapItem] = field(default_factory=list)
    models: list[ModelProfile] = field(default_factory=list)
    landscape: LandscapeSummary | None = None
    download_failures: list[dict[str, Any]] = field(default_factory=list)
    search_log: list[dict[str, Any]] = field(default_factory=list)
    citation_audit: list[CitationAuditRow] = field(default_factory=list)
    citations: dict[str, str] = field(default_factory=dict)
    topic_columns: list[TopicColumn] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------


def _clean(value: Any) -> Any:
    """Coerce any value into something Excel accepts."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (list, tuple, set)):
        return "; ".join(str(v) for v in value if v not in (None, ""))
    text = str(value)
    # Stop Excel interpreting a leading '=' or '+' as a formula.
    if text[:1] in {"=", "+", "-", "@"} and not text.replace(".", "", 1).lstrip("-").isdigit():
        text = "'" + text
    return truncate_text(text, MAX_CELL_CHARS)


def _write_header(sheet: Worksheet, columns: list[dict[str, Any]], row: int = 1) -> None:
    """Write and style one header row."""
    for index, column in enumerate(columns, 1):
        cell = sheet.cell(row=row, column=index, value=column.get("header", ""))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border = CELL_BORDER
        sheet.column_dimensions[get_column_letter(index)].width = float(column.get("width", 20))
    sheet.row_dimensions[row].height = 30


def _finalise(
    sheet: Worksheet,
    columns: list[dict[str, Any]],
    row_count: int,
    *,
    settings: Settings,
    header_row: int = 1,
) -> None:
    """Apply freeze panes, autofilter, and status colours to a finished sheet."""
    reporting = settings.reporting
    if reporting.get("freeze_header", True):
        sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    if reporting.get("autofilter", True) and row_count > 0 and columns:
        last_column = get_column_letter(len(columns))
        sheet.auto_filter.ref = f"A{header_row}:{last_column}{header_row + row_count}"

    colours = settings.status_colours
    if not colours:
        return
    status_indexes = [
        index for index, column in enumerate(columns, 1) if column.get("status_format")
    ]
    for column_index in status_indexes:
        for row_index in range(header_row + 1, header_row + row_count + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            colour = colours.get(str(cell.value or "").lstrip("'"))
            if colour:
                cell.fill = PatternFill("solid", fgColor=colour)


def _write_rows(
    sheet: Worksheet,
    columns: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    *,
    header_row: int = 1,
) -> int:
    """Write data rows, honouring per-column wrap and hyperlink settings."""
    for offset, data in enumerate(rows):
        row_index = header_row + 1 + offset
        for column_index, column in enumerate(columns, 1):
            key = column["key"]
            cell = sheet.cell(row=row_index, column=column_index, value=_clean(data.get(key)))
            cell.font = BODY_FONT
            cell.border = CELL_BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=bool(column.get("wrap")),
                horizontal="left",
            )
            if column.get("hyperlink") and isinstance(cell.value, str):
                url = str(cell.value)
                if url.startswith(("http://", "https://")):
                    cell.hyperlink = url
                    cell.font = LINK_FONT
    return len(rows)


# ---------------------------------------------------------------------------
# Row builders
# ---------------------------------------------------------------------------


def _analysis_value(analysis: PaperAnalysis | None, field_name: str) -> str:
    """Render one analysed field, making its evidence stance explicit."""
    if analysis is None:
        return NOT_REPORTED
    evidence = analysis.field(field_name)
    if not evidence.is_reported:
        return evidence.stance.value if evidence.stance else NOT_REPORTED
    text = evidence.value
    if evidence.stance == EvidenceStance.AGENT_INFERENCE:
        text = f"[Agent inference] {text}"
    if evidence.pages:
        text = f"{text} (p. {', '.join(str(p) for p in evidence.pages)})"
    return text


def _evidence_pages(analysis: PaperAnalysis | None) -> str:
    """Collect every page cited across a paper's analysed fields."""
    if analysis is None:
        return ""
    pages: set[int] = set()
    for name in (
        "research_problem", "research_objective", "study_design", "data_source",
        "model_or_method", "main_findings", "limitations_stated", "gaps_stated_by_authors",
    ):
        pages.update(analysis.field(name).pages)
    return ", ".join(str(p) for p in sorted(pages))


def master_rows(inputs: MatrixInputs) -> list[dict[str, Any]]:
    """Build one row per paper for ``Master Matrix``."""
    rows: list[dict[str, Any]] = []
    for serial, record in enumerate(inputs.records, 1):
        analysis = inputs.analyses.get(record.record_id)
        row: dict[str, Any] = {
            "serial": serial,
            "record_id": record.record_id,
            "title": record.title,
            "authors_joined": "; ".join(record.authors) or NOT_REPORTED,
            "year": record.year or NOT_REPORTED,
            "journal": record.journal or NOT_REPORTED,
            "volume": record.volume or "",
            "issue": record.issue or "",
            "pages": record.pages or record.article_number or "",
            "doi": record.doi or NOT_REPORTED,
            "doi_link": f"https://doi.org/{record.doi}" if record.doi else "",
            "issn": record.issn or "",
            "publisher": record.publisher or "",
            "document_type": record.document_type or "",
            "language": record.language or "",
            "citation_count": record.citation_count if record.citation_count is not None else "",
            "citation_count_retrieved": record.citation_count_retrieved or "",
            "open_access_status": record.open_access_status or "Unknown",
            "licence": record.licence or "",
            "q1.verification_status": record.q1.verification_status.value,
            "q1.quartile": record.q1.quartile or "",
            "q1.ranking_source": record.q1.ranking_source or "None configured",
            "download_status": record.download_status.value,
            "local_filename": record.local_filename or "",
            "file_sha256": (record.file_sha256 or "")[:16],
            "discovery_source": record.discovery_source,
            "metadata_sources_joined": "; ".join(record.metadata_sources),
            "relevance_score": round(record.relevance_score, 3),
            "agent_inferred_gap": (
                analysis.agent_inferred_gap.value if analysis else NOT_REPORTED
            ),
            "evidence_pages": _evidence_pages(analysis),
            "overall_confidence": analysis.overall_confidence if analysis else "Not analysed",
            "verification_confidence": record.verification_confidence,
            "missing_information_joined": (
                "; ".join(analysis.missing_information) if analysis else "Paper not analysed"
            ),
            "notes": record.notes,
        }
        for column in inputs.topic_columns:
            row[f"topic::{column.header}"] = column.values.get(record.record_id, NOT_REPORTED)
        rows.append(row)

    # Analysis-derived columns share one resolution path.
    for serial, record in enumerate(inputs.records, 1):
        analysis = inputs.analyses.get(record.record_id)
        row = rows[serial - 1]
        for key in list(_MASTER_ANALYSIS_KEYS):
            row[f"analysis.{key}"] = _analysis_value(analysis, key)
    return rows


#: Analysis fields shown in ``Master Matrix``.
_MASTER_ANALYSIS_KEYS: tuple[str, ...] = (
    "research_problem", "research_objective", "study_geography", "study_design",
    "data_source", "sample_size", "unit_of_analysis", "dependent_variables",
    "independent_variables", "control_variables", "model_or_method",
    "software_or_tools", "validation_approach", "main_findings",
    "policy_implications", "limitations_stated", "gaps_stated_by_authors",
    "relevance_to_topic",
)


def methods_rows(inputs: MatrixInputs) -> list[dict[str, Any]]:
    """Build ``Methods and Models`` rows."""
    return [
        {
            "serial": serial,
            "model_name": profile.model_name,
            "model_category": profile.model_category,
            "purpose": profile.purpose or NOT_REPORTED,
            "assumptions": profile.assumptions or [NOT_REPORTED],
            "required_inputs": profile.required_inputs or [NOT_REPORTED],
            "outputs": profile.outputs or [NOT_REPORTED],
            "study_application": profile.study_application or [NOT_REPORTED],
            "software_used": profile.software_used or [NOT_REPORTED],
            "calibration_approach": profile.calibration_approach or [NOT_REPORTED],
            "validation_approach": profile.validation_approach or [NOT_REPORTED],
            "advantages": profile.advantages or [NOT_REPORTED],
            "limitations": profile.limitations or [NOT_REPORTED],
            "paper_count": len(profile.paper_record_ids),
            "citations": profile.citations,
            "plain_explanation": profile.plain_explanation,
        }
        for serial, profile in enumerate(inputs.models, 1)
    ]


def findings_rows(inputs: MatrixInputs) -> list[dict[str, Any]]:
    """Build ``Key Findings`` rows."""
    rows: list[dict[str, Any]] = []
    serial = 0
    for record in inputs.records:
        analysis = inputs.analyses.get(record.record_id)
        if analysis is None:
            continue
        findings = analysis.field("main_findings")
        if not findings.is_reported:
            continue
        serial += 1
        rows.append(
            {
                "serial": serial,
                "record_id": record.record_id,
                "in_text_citation": inputs.citations.get(record.record_id, ""),
                "title": record.title,
                "year": record.year or "",
                "study_geography": _analysis_value(analysis, "study_geography"),
                "model_or_method": _analysis_value(analysis, "model_or_method"),
                "main_findings": findings.value,
                "stance": findings.stance.value,
                "pages": ", ".join(str(p) for p in findings.pages),
                "confidence": findings.confidence,
            }
        )
    return rows


def gaps_rows(inputs: MatrixInputs) -> list[dict[str, Any]]:
    """Build ``Research Gaps`` rows."""
    return [
        {
            "serial": serial,
            "gap_id": gap.gap_id,
            "category": gap.category.value,
            "statement": gap.statement,
            "stance": gap.stance.value,
            "citations": gap.citations,
            "evidence_ids": gap.evidence_ids,
        }
        for serial, gap in enumerate(inputs.gaps, 1)
    ]


def landscape_rows(inputs: MatrixInputs) -> list[dict[str, Any]]:
    """Build ``Global Landscape`` rows from the landscape statistics."""
    landscape = inputs.landscape
    if landscape is None:
        return []
    total = max(landscape.total_papers, 1)
    rows: list[dict[str, Any]] = []
    serial = 0
    dimensions = (
        ("Country or city", landscape.countries),
        ("Region", landscape.regions),
        ("Institution", landscape.institutions),
        ("Application", landscape.applications),
        ("Dataset", landscape.datasets),
        ("Method", landscape.methods),
        ("Publication year", landscape.year_counts),
        ("Journal", landscape.journals),
    )
    for name, mapping in dimensions:
        for item, count in sorted(mapping.items(), key=lambda kv: (-kv[1], str(kv[0]))):
            serial += 1
            rows.append(
                {
                    "serial": serial,
                    "dimension": name,
                    "item": item,
                    "paper_count": count,
                    "share": f"{count / total:.0%}",
                    "notes": "Share of the reviewed evidence base, not of world research.",
                }
            )
    for item in landscape.under_researched:
        serial += 1
        rows.append(
            {
                "serial": serial,
                "dimension": "Under-represented",
                "item": item,
                "paper_count": 0,
                "share": "0%",
                "notes": "Not represented in the reviewed evidence; may still exist elsewhere.",
            }
        )
    return rows


def q1_rows(inputs: MatrixInputs) -> list[dict[str, Any]]:
    """Build ``Q1 Verification`` rows."""
    return [
        {
            "serial": serial,
            "record_id": record.record_id,
            "title": record.title,
            "journal_name": record.q1.journal_name or record.journal,
            "issn": record.q1.issn or record.issn or "",
            "eissn": record.q1.eissn or record.eissn or "",
            "publication_year": record.q1.publication_year or record.year or "",
            "ranking_year": record.q1.ranking_year or "Not stated",
            "subject_category": record.q1.subject_category or "Not stated",
            "quartile": record.q1.quartile or "",
            "ranking_source": record.q1.ranking_source or "None configured",
            "verification_date": record.q1.verification_date or "",
            "verification_status": record.q1.verification_status.value,
            "matched_on": record.q1.matched_on or "",
            "notes": record.q1.notes,
        }
        for serial, record in enumerate(inputs.records, 1)
    ]


def citation_audit_rows(inputs: MatrixInputs) -> list[dict[str, Any]]:
    """Build ``Citation Audit`` rows."""
    return [
        {
            "serial": serial,
            "in_text_citation": row.in_text_citation,
            "record_id": row.record_id,
            "doi": row.doi or "",
            "title": row.title,
            "appears_in": row.appears_in_documents or ["not cited"],
            "in_reference_list": row.in_reference_list,
            "title_match": row.title_match.value,
            "author_match": row.author_match.value,
            "year_match": row.year_match.value,
            "journal_match": row.journal_match.value,
            "doi_resolves": row.doi_resolves.value,
            "outcome": row.outcome.value,
            "notes": row.notes,
        }
        for serial, row in enumerate(inputs.citation_audit, 1)
    ]


def data_dictionary_rows(inputs: MatrixInputs, settings: Settings) -> list[dict[str, Any]]:
    """Document every column in every sheet, flagging agent-added columns."""
    rows: list[dict[str, Any]] = []
    definitions = _COLUMN_DEFINITIONS

    for sheet_key in (
        "master_matrix", "methods_and_models", "key_findings", "research_gaps",
        "global_landscape", "download_failures", "search_log", "q1_verification",
        "citation_audit",
    ):
        sheet_name = settings.sheet_name(sheet_key)
        for column in settings.sheet_columns(sheet_key):
            key = column["key"]
            rows.append(
                {
                    "sheet": sheet_name,
                    "column": column.get("header", key),
                    "definition": definitions.get(
                        key,
                        _derive_definition(key),
                    ),
                    "source": _derive_source(key),
                    "added_by_agent": "No",
                    "rationale": "Standard column defined in config/report_columns.yaml.",
                }
            )

    for column in inputs.topic_columns:
        rows.append(
            {
                "sheet": settings.sheet_name("master_matrix"),
                "column": column.header,
                "definition": column.definition,
                "source": "Derived by the analysis agent from the paper text.",
                "added_by_agent": "Yes",
                "rationale": column.rationale,
            }
        )

    rows.append(
        {
            "sheet": "All sheets",
            "column": "Evidence stance values",
            "definition": (
                "'Author explicitly states this' means the wording came from the paper. "
                "'Agent inference based on evidence' means the agent interpreted the text. "
                "'Information not reported' means the paper is silent. "
                "'Information could not be verified' means the source text was unreadable."
            ),
            "source": "Controlled vocabulary in src/literature_review_agent/schemas.py",
            "added_by_agent": "No",
            "rationale": "Ensures inference is never presented as an author statement.",
        }
    )
    rows.append(
        {
            "sheet": settings.sheet_name("q1_verification"),
            "column": "Verification Status values",
            "definition": (
                "Only 'Verified Q1', 'Verified non-Q1', 'Unverified', "
                "'Conflicting information', and 'Not applicable' are used. A quartile is "
                "reported only when a ranking source supplies it."
            ),
            "source": "config/default_config.yaml -> q1_ranking",
            "added_by_agent": "No",
            "rationale": "Prevents a journal's reputation being mistaken for evidence.",
        }
    )
    return rows


#: Definitions for columns whose meaning is not obvious from the header.
_COLUMN_DEFINITIONS: dict[str, str] = {
    "serial": "Sequential row number within this sheet.",
    "record_id": "Stable internal identifier for the paper, used to join sheets.",
    "doi_link": "Resolvable DOI hyperlink for verification.",
    "relevance_score": "0-1 score from title, abstract, keyword, recency, access, and index signals.",
    "file_sha256": "First 16 characters of the downloaded PDF's SHA-256 checksum.",
    "metadata_sources_joined": "Every API whose metadata contributed to this record.",
    "evidence_pages": "Union of page numbers supporting this paper's analysed fields.",
    "overall_confidence": "Confidence in the analysis, from field coverage and readable pages.",
    "verification_confidence": "Confidence after independent verification checks.",
    "missing_information_joined": "Analysis fields the paper did not report.",
    "agent_inferred_gap": "Gap inferred by the agent from what the paper does not report.",
    "q1.verification_status": "One of the five permitted quartile verification states.",
    "q1.ranking_source": "The ranking dataset the quartile came from, or 'None configured'.",
    "download_status": "Outcome of the legal PDF retrieval attempt.",
    "share": "This item's share of the reviewed evidence base, not of world research.",
    "stance": "Whether the authors stated this or the agent inferred it.",
    "attempted_urls": "Every PDF URL tried, including those refused as ineligible.",
    "recommended_action": "How a human can obtain the paper through their own access.",
}


def _derive_definition(key: str) -> str:
    """Generate a readable definition for a straightforward column."""
    readable = key.split(".")[-1].replace("_", " ")
    if key.startswith("analysis."):
        return (
            f"The paper's {readable}, as extracted from its text with page evidence. "
            "Prefixed '[Agent inference]' when interpreted rather than stated."
        )
    return f"The paper's {readable} as recorded in the bibliographic metadata."


def _derive_source(key: str) -> str:
    """Say where a column's values come from."""
    if key.startswith("analysis."):
        return "Paper full text (analysis stage)"
    if key.startswith("q1"):
        return "Journal ranking file (Q1 verification stage)"
    if key in {"download_status", "local_filename", "file_sha256", "failure_reason", "http_status"}:
        return "Download stage"
    if key in {"relevance_score", "discovery_source", "metadata_sources_joined"}:
        return "Discovery and scoring stages"
    return "Scholarly API metadata"


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------


def build_workbook(inputs: MatrixInputs, settings: Settings, output_path: Path) -> Path:
    """Write the complete ten-sheet literature-review matrix."""
    output_path = Path(output_path)
    ensure_dir(output_path.parent)

    max_topic = int(settings.reporting.get("max_topic_specific_columns", 10))
    if len(inputs.topic_columns) > max_topic:
        LOG.warning(
            f"{len(inputs.topic_columns)} topic-specific columns were proposed but the "
            f"limit is {max_topic}; keeping the first {max_topic}."
        )
        inputs.topic_columns = inputs.topic_columns[:max_topic]

    workbook = Workbook()
    workbook.remove(workbook.active)

    # --- 1. Master Matrix ---
    columns = settings.sheet_columns("master_matrix")
    for column in inputs.topic_columns:
        columns.append(
            {
                "key": f"topic::{column.header}",
                "header": column.header,
                "width": column.width,
                "wrap": True,
            }
        )
    sheet = workbook.create_sheet(settings.sheet_name("master_matrix"))
    _write_header(sheet, columns)
    count = _write_rows(sheet, columns, master_rows(inputs))
    _finalise(sheet, columns, count, settings=settings)

    # --- 2-9. remaining data sheets ---
    for sheet_key, builder in (
        ("methods_and_models", methods_rows),
        ("key_findings", findings_rows),
        ("research_gaps", gaps_rows),
        ("global_landscape", landscape_rows),
        ("q1_verification", q1_rows),
        ("citation_audit", citation_audit_rows),
    ):
        sheet_columns = settings.sheet_columns(sheet_key)
        sheet = workbook.create_sheet(settings.sheet_name(sheet_key))
        _write_header(sheet, sheet_columns)
        rows = builder(inputs)
        written = _write_rows(sheet, sheet_columns, rows)
        _finalise(sheet, sheet_columns, written, settings=settings)
        if not rows:
            _write_empty_note(sheet, len(sheet_columns), sheet_key)

    # --- Download Failures ---
    failure_columns = settings.sheet_columns("download_failures")
    sheet = workbook.create_sheet(settings.sheet_name("download_failures"))
    _write_header(sheet, failure_columns)
    written = _write_rows(sheet, failure_columns, inputs.download_failures)
    _finalise(sheet, failure_columns, written, settings=settings)
    if not inputs.download_failures:
        _write_empty_note(sheet, len(failure_columns), "download_failures")

    # --- Search Log ---
    log_columns = settings.sheet_columns("search_log")
    sheet = workbook.create_sheet(settings.sheet_name("search_log"))
    _write_header(sheet, log_columns)
    log_rows = [
        {"serial": index, **{k: entry.get(k, "") for k in
                             ("timestamp", "source", "query", "breadth", "results",
                              "http_status", "outcome", "notes")}}
        for index, entry in enumerate(inputs.search_log, 1)
    ]
    written = _write_rows(sheet, log_columns, log_rows)
    _finalise(sheet, log_columns, written, settings=settings)

    # --- 10. Data Dictionary ---
    dictionary_columns = settings.sheet_columns("data_dictionary")
    sheet = workbook.create_sheet(settings.sheet_name("data_dictionary"))
    _write_header(sheet, dictionary_columns)
    written = _write_rows(sheet, dictionary_columns, data_dictionary_rows(inputs, settings))
    _finalise(sheet, dictionary_columns, written, settings=settings)

    workbook.save(output_path)
    LOG.info(
        f"Wrote {output_path.name} with {len(workbook.sheetnames)} sheets "
        f"({len(inputs.records)} papers)."
    )
    return output_path


def _write_empty_note(sheet: Worksheet, column_count: int, sheet_key: str) -> None:
    """Explain an empty sheet instead of leaving a reviewer guessing.

    Written into the first data cell only — never a merged cell, so filters and
    sorting keep working.
    """
    notes = {
        "methods_and_models": (
            "No models or methods were identified. This happens when no PDF could be "
            "analysed, or when the analysed papers do not name a recognised method."
        ),
        "key_findings": (
            "No paper reported findings that could be extracted with page evidence."
        ),
        "research_gaps": "No research gaps were identified from the analysed papers.",
        "global_landscape": (
            "No landscape statistics could be computed, because no paper was analysed."
        ),
        "download_failures": (
            "No download failures: every selected paper with a legal open-access URL "
            "was retrieved and validated."
        ),
        "citation_audit": "No citations were generated, so there was nothing to audit.",
        "q1_verification": "No papers were verified.",
    }
    cell = sheet.cell(row=2, column=1, value=notes.get(sheet_key, "No rows for this sheet."))
    cell.font = Font(size=10, italic=True, color="808080")
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_evidence_workbook(
    rows: list[dict[str, Any]],
    output_path: Path,
    *,
    title: str = "Evidence Ledger",
) -> Path:
    """Write ``Evidence_Ledger.xlsx``: every claim with its source evidence."""
    output_path = Path(output_path)
    ensure_dir(output_path.parent)

    columns = [
        {"key": "serial", "header": "S. No.", "width": 7},
        {"key": "evidence_id", "header": "Evidence ID", "width": 18},
        {"key": "document", "header": "Document", "width": 26},
        {"key": "section", "header": "Section", "width": 24, "wrap": True},
        {"key": "claim", "header": "Claim", "width": 60, "wrap": True},
        {"key": "in_text_citation", "header": "In-Text Citation", "width": 24},
        {"key": "record_id", "header": "Record ID", "width": 14},
        {"key": "doi", "header": "DOI", "width": 24},
        {"key": "field_name", "header": "Analysis Field", "width": 22},
        {"key": "stance", "header": "Evidence Stance", "width": 28, "wrap": True},
        {"key": "pages", "header": "Pages", "width": 12},
        {"key": "supporting_text", "header": "Supporting Text", "width": 60, "wrap": True},
        {"key": "confidence", "header": "Confidence", "width": 12},
        {"key": "created_at", "header": "Recorded At", "width": 20},
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]
    _write_header(sheet, columns)
    for offset, data in enumerate(rows):
        row_index = 2 + offset
        for column_index, column in enumerate(columns, 1):
            cell = sheet.cell(
                row=row_index, column=column_index, value=_clean(data.get(column["key"]))
            )
            cell.font = BODY_FONT
            cell.border = CELL_BORDER
            cell.alignment = Alignment(
                vertical="top", wrap_text=bool(column.get("wrap")), horizontal="left"
            )
    sheet.freeze_panes = "A2"
    if rows:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"
    else:
        _write_empty_note(sheet, len(columns), "evidence")

    workbook.save(output_path)
    LOG.info(f"Wrote {output_path.name} with {len(rows)} evidence record(s).")
    return output_path


def build_citation_audit_workbook(
    rows: list[dict[str, Any]], output_path: Path, settings: Settings
) -> Path:
    """Write the standalone ``Citation_Audit.xlsx``."""
    output_path = Path(output_path)
    ensure_dir(output_path.parent)
    columns = settings.sheet_columns("citation_audit")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Citation Audit"
    _write_header(sheet, columns)
    written = _write_rows(sheet, columns, rows)
    _finalise(sheet, columns, written, settings=settings)
    if not rows:
        _write_empty_note(sheet, len(columns), "citation_audit")
    workbook.save(output_path)
    LOG.info(f"Wrote {output_path.name} with {len(rows)} audit row(s).")
    return output_path
