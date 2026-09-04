"""Build notebooks/literature_review_pipeline.ipynb with nbformat."""
from pathlib import Path
import nbformat as nbf

nb = nbf.v4.new_notebook()
C = []
def md(text): C.append(nbf.v4.new_markdown_cell(text.strip("\n")))
def code(text): C.append(nbf.v4.new_code_cell(text.strip("\n")))

md("""
# Literature Review Pipeline

A resumable, verification-first literature review, run step by step.

**Edit the inputs in section 2, then run the cells in order.** Every cell calls
functions from the `literature_review_agent` package — this notebook holds no
research logic of its own, so anything you can do here you can also do from the
command line, and vice versa.

**Where things go.** The five numbered folders on disk are staging. Outputs are
uploaded to Google Drive and verified against a real file ID before being
counted as delivered. If Drive is not configured, the notebook says so plainly
and leaves everything in local staging — it never claims an upload it cannot
prove.

**Interruptions are safe.** Every stage checkpoints. If your connection drops or
you interrupt the kernel, re-run the cell: completed work is skipped and long
stages resume at the item they stopped on.
""")

md("## 1. Environment check\n\nConfirms the package imports, locates the project root, and reports which\nsources and credentials are available. No network calls are made here.")

code('''
import sys
from pathlib import Path

# Locate the project root without hard-coding any absolute path: walk up from
# the notebook until the directory holding pyproject.toml / config/ is found.
def find_project_root(start: Path) -> Path:
    """Return the first ancestor containing the project's marker files."""
    for candidate in (start, *start.parents):
        if (candidate / "config" / "default_config.yaml").exists():
            return candidate
    raise RuntimeError(
        "Could not locate the project root. Run this notebook from inside the "
        "'Literature Review' folder."
    )

PROJECT_ROOT = find_project_root(Path.cwd().resolve())
SRC = PROJECT_ROOT / "src"
if SRC.exists() and str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

print(f"Project root: {PROJECT_ROOT}")
print(f"Python: {sys.version.split()[0]}")
''')

code('''
import literature_review_agent as lra
from literature_review_agent.config import available_keys, load_settings, missing_keys
from literature_review_agent.drive_storage import describe_drive_readiness

SETTINGS = load_settings(PROJECT_ROOT)
SETTINGS.ensure_top_level_dirs()

print(f"literature_review_agent version {lra.__version__}\\n")

print("Search sources")
for name, spec in SETTINGS.source_specs().items():
    state = "available" if spec.available else f"skipped ({spec.unavailable_reason})"
    print(f"  {spec.label:38s} {state}")

print("\\nOptional credentials")
for name, present in available_keys().items():
    print(f"  {name:34s} {'set' if present else 'not set'}")

print("\\nGoogle Drive")
readiness = describe_drive_readiness(SETTINGS)
print(f"  {readiness.summary()}")
if readiness.setup_steps:
    print("\\n  To finish Drive setup:")
    for index, step in enumerate(readiness.setup_steps, 1):
        print(f"    {index}. {step}")

ranking = SETTINGS.q1_ranking.get("file")
resolved = SETTINGS.resolve_path(str(ranking)) if ranking else None
print("\\nJournal-ranking data")
if resolved and resolved.exists():
    print(f"  configured: {resolved.name}")
else:
    print("  not configured - every quartile will be reported as 'Unverified'.")
    print("  No quartile is ever guessed. Supply a licensed Scimago or JCR")
    print("  export and set q1_ranking.file in config/default_config.yaml.")
''')

md("""
## 2. Your inputs

**This is the only cell you need to edit.**

Leave a value as-is to accept the documented default. Whatever you do not set is
recorded as an assumption in `job_config.yaml` and repeated in the final summary,
so nothing is silently decided for you.
""")

code('''
# --- Required -------------------------------------------------------------
TOPIC = "Effect of rainfall on urban travel behaviour"

# --- Strongly recommended ------------------------------------------------
RESEARCH_QUESTIONS = [
    "How does rainfall intensity influence mode choice and daily travel?",
]

# --- Scope ---------------------------------------------------------------
YEAR_FROM = 2015
YEAR_TO = 2026
MAX_PAPERS = 50
Q1_MODE = "preferred"          # "only" | "preferred" | "ignore"
GEOGRAPHY = "global"           # e.g. "India", "South Asia", "global"
LANGUAGE = "English"
PAPER_TYPES = ["journal article"]
CITATION_STYLE = "APA 7"       # APA 7 | Harvard | IEEE | Vancouver | Chicago 17

# --- Optional refinements ------------------------------------------------
USER_KEYWORDS: list[str] = []      # terms you want searched, e.g. ["mode choice"]
EXCLUSION_TERMS: list[str] = []    # terms that mark a record irrelevant
RANKING_FILE = None                # e.g. "config/scimago_2024.csv"
OUTPUT_ROOT = "."                  # where the numbered folders live

# --- Storage -------------------------------------------------------------
# None  = follow config/google_drive.yaml
# True  = require Drive uploads
# False = stay local for this run
USE_GOOGLE_DRIVE = None
''')

md("## 3. Create the job\n\nCreates the date and topic partitioned folders and records the configuration,\nincluding the complete original topic and every assumption applied.")

code('''
from literature_review_agent.job_manager import Job

overrides = {
    "research_questions": RESEARCH_QUESTIONS,
    "year_from": YEAR_FROM,
    "year_to": YEAR_TO,
    "maximum_papers": MAX_PAPERS,
    "q1_mode": Q1_MODE,
    "geography": GEOGRAPHY,
    "language": LANGUAGE,
    "paper_types": PAPER_TYPES,
    "citation_style": CITATION_STYLE,
    "user_keywords": USER_KEYWORDS,
    "exclusion_terms": EXCLUSION_TERMS,
    "ranking_file": RANKING_FILE,
    "output_root": OUTPUT_ROOT,
}
overrides = {k: v for k, v in overrides.items() if v not in (None, [], "")}

JOB = Job.create(TOPIC, settings=SETTINGS, **overrides)

print(f"Topic:      {JOB.config.topic}")
print(f"Job folder: {JOB.paths.logs}")
print(f"Date:       {JOB.config.job_date}")
print(f"Slug:       {JOB.config.topic_slug}")
print(f"Scope:      {JOB.config.year_from}-{JOB.config.year_to}, "
      f"max {JOB.config.maximum_papers} papers, Q1 mode '{JOB.config.q1_mode.value}'")

print("\\nAssumptions recorded:")
for assumption in JOB.config.assumptions:
    print(f"  - {assumption}")
''')

code('''
from literature_review_agent.orchestrator import Orchestrator

ORCH = Orchestrator(JOB, settings=SETTINGS, enable_drive=USE_GOOGLE_DRIVE)
ORCH.load_state()

# Create the Drive folder tree up front, so the destinations exist before any
# artefact is produced. Returns None when Drive is not usable, in which case the
# run continues locally and says so.
tree = ORCH.storage.ensure_remote_tree()
if tree:
    print(f"Google Drive tree ready ({len(tree)} folders).")
else:
    print("Google Drive is not active: outputs stay in local staging.")
    print(f"  Reason: {ORCH.storage.readiness.summary()}")
''')

md("""
### A helper for running stages

Prints the outcome of each stage and stops at the first failure, leaving the
checkpoint intact so you can fix the cause and re-run the same cell.
""")

code('''
from literature_review_agent.schemas import STAGE_ORDER, StageName, StageStatus


def run(*stages: StageName, force: bool = False) -> None:
    """Run the given stages, reporting each outcome."""
    for stage in stages:
        outcome = ORCH.run_stage(stage, force=force)
        marker = {
            StageStatus.COMPLETE: "done",
            StageStatus.SKIPPED: "skipped",
            StageStatus.FAILED: "FAILED",
        }.get(outcome.status, outcome.status.value)
        print(f"[{marker}] {stage.value}: {outcome.message}")
        if outcome.counters:
            for key, value in outcome.counters.items():
                print(f"         {key}: {value}")
        if outcome.status is StageStatus.FAILED:
            print("\\nProgress is checkpointed. Fix the cause and re-run this cell.")
            break


def show_status() -> None:
    """Print the stage-by-stage state of the job."""
    description = JOB.describe()
    for stage in STAGE_ORDER:
        info = description["stages"].get(stage.value, {})
        print(f"  {stage.value:16s} {info.get('status', 'pending'):10s} "
              f"items={info.get('items_done', 0):3d}  {info.get('message', '')[:60]}")
    print(f"\\n  next stage: {description['next_stage'] or 'none - job complete'}")
''')

md("""
## 4. Keyword generation

Produces main concepts, synonyms, abbreviations, alternative spellings, related
methods, application terms, geographic terms, exclusion terms, and
database-ready Boolean strings for Scopus, Web of Science, Crossref, OpenAlex,
Semantic Scholar and, where relevant, IEEE Xplore, PubMed/Europe PMC and TRID.

Every term records whether you supplied it or the agent generated it.
""")

code("run(StageName.KEYWORDS)")

code('''
import pandas as pd

terms = pd.read_csv(JOB.paths.keywords / "keywords.csv")
print(f"{len(terms)} terms\\n")
print(terms.groupby(["category", "provenance"]).size().to_string())

print("\\nMain concepts:")
for concept in (ORCH.state.strategy.main_concepts if ORCH.state.strategy else []):
    print(f"  - {concept}")

print("\\nRecommended search string (balanced):")
for item in (ORCH.state.strategy.search_strings if ORCH.state.strategy else []):
    if item.database == "generic" and item.breadth == "balanced":
        print(f"  {item.query}")
''')

md("""
## 5. Search

Queries every available source, records one log line per query, and screens
records against the year range, document type, and your exclusion terms.

Crossref, OpenAlex, Semantic Scholar, Europe PMC, and arXiv need no
credentials. CORE, Elsevier, and Springer are skipped without their API keys —
a coverage limitation, not a failure.
""")

code("run(StageName.SEARCH)")

md("## 6. Metadata deduplication\n\nMerges duplicates in a fixed order — DOI, exact title, fuzzy title corroborated\nby author and year, then other identifiers — enriches gaps from Crossref and\nUnpaywall, and scores relevance. Merges are additive: the richer record\nsurvives and absorbs the other's useful fields, with an audit trail.")

code("run(StageName.DEDUPLICATE)")

code('''
records = ORCH.state.records
print(f"{len(records)} unique records\\n")

if records:
    frame = pd.DataFrame([
        {
            "title": r.title[:60],
            "year": r.year,
            "journal": (r.journal or "")[:34],
            "doi": r.doi,
            "oa": r.open_access_status,
            "relevance": round(r.relevance_score, 3),
            "sources": len(r.metadata_sources),
        }
        for r in sorted(records, key=lambda r: -r.relevance_score)
    ])
    display(frame.head(20))

    merged = sum(1 for r in records if r.merged_from)
    print(f"\\n{merged} record(s) absorbed a duplicate.")
''')

md("""
## 7. Q1 journal verification

Quartiles come **only** from a licensed ranking file you supply. This project
ships none, because that data is licensed.

Without a file, every paper is reported `Unverified`. No journal is treated as
Q1 because it is famous, highly cited, or published by a major publisher. A
journal holding different quartiles across subject categories is reported as
`Conflicting information` for you to resolve, not silently decided.
""")

code("run(StageName.Q1_VERIFY, StageName.SELECT)")

code('''
from collections import Counter

status_counts = Counter(r.q1.verification_status.value for r in ORCH.state.records)
for status, count in status_counts.most_common():
    print(f"  {status:28s} {count}")

selected = [r for r in ORCH.state.records if r.selected]
print(f"\\n{len(selected)} paper(s) selected for the review.")

pending = JOB.paths.verification / "pending_q1_verification.csv"
if pending.exists():
    print(f"\\nPapers awaiting manual quartile verification: {pending}")
    display(pd.read_csv(pending)[["title", "journal", "verification_status"]].head(10))
''')

md("""
## 8. PDF download

Retrieves each selected paper's PDF **only** from a legitimate open-access
location, or from a direct PDF URL a publisher's own API has identified as
authorised.

Never bypassed: paywalls, institutional logins, CAPTCHAs, anti-bot protections.
Never contacted: Sci-Hub and comparable mirrors, or Google Scholar.

Each file is validated before it is accepted — PDF signature, parseability,
page count, and rejection of HTML error pages saved as `.pdf` — then renamed to
its paper title with a recorded checksum. Papers that cannot be obtained
legally go to `Unable_to_Download.docx` with a recommended manual action.
""")

code("run(StageName.DOWNLOAD)")

code('''
from literature_review_agent.schemas import DownloadStatus

downloaded = [r for r in ORCH.state.records
              if r.download_status in (DownloadStatus.DOWNLOADED,
                                       DownloadStatus.ALREADY_PRESENT)]
failed = [r for r in ORCH.state.records
          if r.download_status in (DownloadStatus.FAILED,
                                   DownloadStatus.SKIPPED_NO_LEGAL_URL)]

print(f"Downloaded and validated: {len(downloaded)}")
for record in downloaded:
    size = (record.file_bytes or 0) / 1024
    print(f"  {record.local_filename}  ({size:,.0f} KB)")

if failed:
    print(f"\\nCould not be obtained legally: {len(failed)}")
    for record in failed:
        print(f"  {record.title[:66]}")
        print(f"    reason: {record.failure_reason}")
    print(f"\\nFull register with manual actions:")
    print(f"  {JOB.paths.unable_to_download / 'Unable_to_Download.docx'}")
''')

md("## 9. PDF text extraction\n\nExtracts text with `=== PAGE n ===` markers preserved, so every later claim can\ncite the page it came from. Scanned PDFs with no text layer are detected and\nflagged as needing OCR — **no text is invented to fill them**, and they support\nno evidence-based claim.")

code("run(StageName.EXTRACT)")

code('''
for record in ORCH.state.records:
    if not record.extracted_text_path:
        continue
    flag = "  [needs OCR]" if record.requires_ocr else ""
    print(f"  {record.title[:56]:58s} {record.extracted_pages or 0:3d} pages, "
          f"{record.extracted_characters or 0:7,d} chars{flag}")

needs_ocr = [r for r in ORCH.state.records if r.requires_ocr]
if needs_ocr:
    print(f"\\n{len(needs_ocr)} PDF(s) have no text layer. They are flagged rather")
    print("than guessed at, and contribute no evidence to the reports.")
''')

md("""
## 10. Paper analysis

Reads each paper's saved text and records, field by field, what it says — with
page numbers.

Every field carries one of four states, and the distinction reaches the final
documents:

- `Author explicitly states this`
- `Agent inference based on evidence`
- `Information not reported`
- `Information could not be verified`

A detail is never inferred merely because a method commonly implies it.
""")

code("run(StageName.ANALYSE, StageName.VERIFY_EVIDENCE)")

code('''
from literature_review_agent.schemas import ANALYSIS_FIELDS, EvidenceStance

analyses = ORCH.state.analyses
print(f"{len(analyses)} paper(s) analysed\\n")

if analyses:
    coverage = pd.DataFrame([
        {
            "field": name.replace("_", " "),
            "reported": sum(1 for a in analyses.values() if a.field(name).is_reported),
            "author-stated": sum(
                1 for a in analyses.values()
                if a.field(name).stance == EvidenceStance.AUTHOR_STATED
                and a.field(name).is_reported
            ),
        }
        for name in ANALYSIS_FIELDS
    ])
    coverage["of total"] = coverage["reported"].astype(str) + f" / {len(analyses)}"
    display(coverage[["field", "of total", "author-stated"]])

    print(f"\\nEvidence ledger: {len(ORCH.state.ledger.records)} claim(s) recorded")
    for stance, count in ORCH.state.ledger.counts_by_stance().items():
        print(f"  {stance:34s} {count}")
''')

md("""
## 11. Literature review matrix

Builds `Literature_Review_Matrix.xlsx` with one row per paper and ten sheets:
Master Matrix, Methods and Models, Key Findings, Research Gaps, Global
Landscape, Download Failures, Search Log, Q1 Verification, Citation Audit, and a
Data Dictionary documenting every column — including any topic-specific column
the agent added, with its rationale.
""")

md("""
## 12. Word reports

Writes the five synthesis documents. Each carries a title, the topic, the date,
its scope, the evidence base, in-text citations, a reference list, limitations,
and a verification note. Every substantive statement maps to an evidence-ledger
record; statistics describe the reviewed evidence, never world research.
""")

code("run(StageName.REPORT)")

code('''
for path in sorted(JOB.paths.reports.glob("*")):
    print(f"  {path.name:42s} {path.stat().st_size / 1024:8,.1f} KB")

matrix = JOB.paths.reports / SETTINGS.reporting.get(
    "excel_filename", "Literature_Review_Matrix.xlsx"
)
if matrix.exists():
    from openpyxl import load_workbook
    book = load_workbook(matrix, read_only=True)
    print(f"\\n{matrix.name} sheets:")
    for name in book.sheetnames:
        print(f"  {name}")
    book.close()
''')

md("""
## 13. Verification

Three independent verifiers run over the saved artefacts. The agent that
analysed the papers does not certify its own work.

- **Metadata and Q1**: DOI resolution; title, author, year, journal and ISSN
  agreement; whether every quartile claim has a dated source; whether duplicate
  merges were sound.
- **Evidence and citations**: whether every claim is supported; whether page
  numbers and quoted text are real; whether numeric values match the source;
  whether citations and reference lists correspond in both directions; whether
  author-stated and agent-inferred content stayed separate.
- **Files**: whether each PDF is valid, is the intended paper, matches its
  checksum, and is correctly paired with its extracted text.

Questionable data is never silently rewritten: the original value is preserved
beside any correction, with its source and reason.
""")

code("run(StageName.FINAL_VERIFY)")

code('''
from literature_review_agent.utils import read_json

findings = read_json(JOB.paths.findings_file, []) or []
outcomes = Counter(f.get("outcome") for f in findings)
print(f"{len(findings)} check(s) run")
for outcome, count in outcomes.most_common():
    print(f"  {outcome:18s} {count}")

failures = [f for f in findings if f.get("outcome") == "Fail"]
if failures:
    print(f"\\n{len(failures)} unresolved problem(s):")
    for finding in failures[:12]:
        print(f"  [{finding['check']}] {finding['detail'][:96]}")
    print("\\nThe review is not complete while these stand. Full detail:")
    print(f"  {JOB.paths.verification / 'unresolved_issues.csv'}")
else:
    print("\\nNo check failed.")
''')

md("## 14. Final output summary\n\nWhat was produced, what was assumed, what remains outstanding, and the exact\nGoogle Drive upload state read from the manifest — never an assumed one.")

code('''
summary = ORCH.completion_summary()

print(f"Topic:      {summary['topic']}")
print(f"Job folder: {summary['job_folder']}")
print(f"Complete:   {summary['complete']}\\n")

print("Counts")
for label, key in (
    ("papers discovered", "discovered"),
    ("unique after dedup", "unique"),
    ("included in review", "included"),
    ("verified Q1", "verified_q1"),
    ("unverified quartile", "unverified_quartile"),
    ("PDFs downloaded", "downloaded"),
    ("failed downloads", "failed_downloads"),
    ("papers analysed", "analysed"),
    ("evidence records", "evidence_records"),
    ("unresolved problems", "unresolved_problems"),
):
    print(f"  {label:22s} {summary['counts'].get(key, 0)}")

print("\\nGoogle Drive")
drive = summary["drive"]
if drive["drive_enabled"]:
    print(f"  {drive['artefacts_uploaded_and_verified']} of "
          f"{drive['artefacts_tracked']} artefacts uploaded and verified")
    if drive["artefacts_pending_upload"]:
        print(f"  {drive['artefacts_pending_upload']} pending. Retry with:")
        print(f'    python -m literature_review_agent drive-sync --job "{JOB.paths.logs}"')
    for link in drive["verified_links"][:8]:
        print(f"    {link['filename']}: {link['link']}")
else:
    print(f"  not active - {drive['drive_status']}")
    print("  All outputs are in local staging only. Nothing has been uploaded.")

print("\\nAssumptions applied")
for assumption in summary["assumptions"]:
    print(f"  - {assumption}")

if summary["limitations"]:
    print("\\nLimitations and outstanding items")
    for limitation in summary["limitations"]:
        print(f"  - {limitation}")

if summary["missing_api_keys"]:
    print(f"\\nOptional credentials not set: {', '.join(summary['missing_api_keys'])}")

print("\\nStage state")
show_status()
''')

md("""
## Resuming later

The job lives entirely on disk, so you can close this notebook and come back.

From the command line:

```bash
python -m literature_review_agent status --job "05 Logs and State/<date>/<slug>"
python -m literature_review_agent resume --job "05 Logs and State/<date>/<slug>"
```

Or in this notebook, re-run section 3 with the same `TOPIC` — `Job.create`
reuses the existing folder — then run the remaining stage cells. Completed
stages are skipped automatically.

To re-run a stage deliberately, pass `force=True`:

```python
run(StageName.REPORT, force=True)
```

### Loading an earlier job

```python
from literature_review_agent.job_manager import Job, list_jobs

for entry in list_jobs(SETTINGS):
    print(entry["job_date"], entry["stages_complete"], entry["topic"])

JOB = Job.load("05 Logs and State/<date>/<slug>", settings=SETTINGS)
ORCH = Orchestrator(JOB, settings=SETTINGS, enable_drive=USE_GOOGLE_DRIVE)
ORCH.load_state()
```
""")

nb["cells"] = C
nb["metadata"] = {
    "kernelspec": {"display_name": "Python 3", "language": "python", "name": "python3"},
    "language_info": {"name": "python", "version": "3.11.0",
                      "mimetype": "text/x-python",
                      "file_extension": ".py",
                      "pygments_lexer": "ipython3",
                      "nbconvert_exporter": "python",
                      "codemirror_mode": {"name": "ipython", "version": 3}},
}
out = Path("notebooks/literature_review_pipeline.ipynb")
out.parent.mkdir(parents=True, exist_ok=True)
nbf.write(nb, out)
print(f"wrote {out} with {len(C)} cells "
      f"({sum(1 for c in C if c.cell_type=='markdown')} markdown, "
      f"{sum(1 for c in C if c.cell_type=='code')} code)")
